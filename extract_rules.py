import openpyxl
import json

wb = openpyxl.load_workbook('Reglas_de_auditoria_automatizada_20251216.xlsx')

rules = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"HOJA: {sheet_name}")
    print('='*80)
    
    # Extract first 40 rows to capture the rule details
    for row_idx, row in enumerate(sheet.iter_rows(max_row=40), 1):
        row_data = []
        for cell in row:
            value = cell.value if cell.value is not None else ""
            row_data.append(str(value))
        
        # Only print non-empty rows
        if any(val.strip() for val in row_data):
            print(f"Fila {row_idx}: {' | '.join(row_data)}")
    
    print("\n")
