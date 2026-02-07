import openpyxl
import json
import sys

# Set UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Reglas_de_auditoria_automatizada_20251216.xlsx')

all_rules = {}

print(f"Total de hojas encontradas: {len(wb.sheetnames)}\n")
print("="*100)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    print(f"\n{'='*100}")
    print(f"HOJA: {sheet_name}")
    print('='*100)
    
    rows_data = []
    
    # Extract all rows (increase max_row to capture more content)
    for row_idx, row in enumerate(sheet.iter_rows(max_row=100), 1):
        row_values = []
        has_content = False
        
        for cell in row:
            value = cell.value if cell.value is not None else ""
            row_values.append(str(value))
            if str(value).strip():
                has_content = True
        
        # Only store and print non-empty rows
        if has_content:
            rows_data.append(row_values)
            # Print with row number
            print(f"Fila {row_idx:3d}: {' | '.join(row_values)}")
    
    all_rules[sheet_name] = rows_data
    print(f"\nTotal de filas con contenido: {len(rows_data)}")

# Save to JSON for later processing
with open('reglas_extraidas.json', 'w', encoding='utf-8') as f:
    json.dump(all_rules, f, ensure_ascii=False, indent=2)

print("\n" + "="*100)
print(f"Extracción completa. Total de hojas procesadas: {len(all_rules)}")
print("Datos guardados en: reglas_extraidas.json")
